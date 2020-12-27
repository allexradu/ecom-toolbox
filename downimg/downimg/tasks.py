from celery import shared_task
from time import sleep
from celery_progress.backend import ProgressRecorder
from django.http import HttpResponse
import openpyxl as xl
import requests
import os
import string
import random
import shutil
from .settings import MEDIA_URL
import platform
from django.core.files.storage import FileSystemStorage
from socket import error as SocketError
from urllib.error import URLError, HTTPError
from . import delete

write_col_start = ''


@shared_task(bind = True)
def go_to_sleep(self, duration, xl_file_name):
    progress_recorder = ProgressRecorder(self)
    for i in range(5):
        sleep(duration)
        progress_recorder.set_progress(i + 1, 5, f'On iteration {i}')
    file_url = '/media/' + xl_file_name
    return HttpResponse(f'<a href="{file_url}">Download</a>')


def clean_slashes(path_elem):
    elem = path_elem.replace('/', '') if path_elem.find('/') != - 1 else path_elem
    elem = elem.replace('\\', '') if elem.find('/') != -1 else elem
    return elem


def platform_combine(*args, is_file):
    """This function combines multiple path elements use when os.path.join doesn't work """
    
    def get_platform_slash():
        return '\\' if platform.system() == 'Windows' else '/'
    
    platform_path = ''
    
    if is_file:
        for i in range(len(args)):
            if i != len(args) - 1:
                platform_path = platform_path + args[i] + get_platform_slash()
            else:
                platform_path = platform_path + args[i]
    else:
        for arg in args:
            platform_path = platform_path + arg + get_platform_slash()
    return platform_path


def get_paths(xl_file_name, folder_name):
    cwd = os.path.abspath(os.curdir)
    table_location = platform_combine(cwd, clean_slashes(MEDIA_URL), xl_file_name, is_file = True)
    media_path = platform_combine(cwd, clean_slashes(MEDIA_URL), is_file = False)
    dir_name = ''.join(random.choices(string.ascii_letters + string.digits, k = 16))
    dir_path = platform_combine(cwd, clean_slashes(MEDIA_URL), dir_name, is_file = False)
    folder = os.path.join(dir_path, folder_name)
    archive_file_name = dir_name + '.zip'
    archive_path = platform_combine(cwd, clean_slashes(MEDIA_URL), dir_name, is_file = True)
    file_url = MEDIA_URL + archive_file_name
    return cwd, table_location, media_path, dir_name, dir_path, folder, archive_file_name, archive_path, file_url


def archive_and_save(d_archive_path, d_dir_path, d_file_url):
    shutil.make_archive(d_archive_path, 'zip', d_dir_path)
    
    fs = FileSystemStorage()
    fs.url(d_file_url)
    try:
        shutil.rmtree(d_dir_path)
        return True
    except OSError as e:
        return f'Error deleting folder: {d_dir_path}  error: {e.strerror}'


@shared_task(bind = True)
def process_file(self, url_cell_start, url_cell_end, write_cell_start, write_cell_end, xl_file_name):
    (cwd, table_location, media_path, dir_name, dir_path, img_folder, archive_file_name, archive_path,
     file_url) = get_paths(xl_file_name, 'photos')
    
    delete.delete_file_and_folders(media_path, 1)
    
    progress_recorder = ProgressRecorder(self)
    
    wb = xl.load_workbook(table_location)
    sh = wb[wb.sheetnames[0]]
    
    os.makedirs(img_folder)
    
    global write_col_start
    write_col_start = write_cell_start
    
    def update_excel(wr_cell_start, wr_cell_end, u_cell_start, img_f_name, t_row, t_col):
        """Updating the excel file with the image file name"""
        global write_col_start
        diff = wr_cell_start - u_cell_start
        sh.cell(t_row, t_col + diff).value = img_f_name
    
    for row in range(2, sh.max_row + 1):
        for col in range(url_cell_start, url_cell_end + 1):
            def get_response():
                """Getting the response from the url get request and returning or returning None if error"""
                try:
                    resp = requests.get(img_url, stream = True, verify = False)
                    return resp
                except HTTPError or URLError or SocketError:
                    return None
            
            img_url = sh.cell(row, col).value
            
            if img_url is not None:
                img_file_name = img_url.split('/')[-1]
                img_file_path = os.path.join(img_folder, img_file_name)
                response = get_response()
                
                if response is not None:
                    with open(img_file_path, 'wb') as out_file:
                        shutil.copyfileobj(response.raw, out_file)
                        progress_recorder.set_progress(row, sh.max_row + 1,
                                                       f'Downloading Row: {row} Column: {col} Image: {img_file_name}')
                        print('downloading image: ', img_url + ' row: ' + str(row) + ' col: ' + str(col))
                    
                    update_excel(write_cell_start, write_cell_end, url_cell_start, img_file_name, row, col)
                
                else:
                    update_excel(write_cell_start, write_cell_end, url_cell_start, 'n/a', row, col)
                    
                    progress_recorder.set_progress(row, sh.max_row + 1,
                                                   f'ERROR on downloading Column: {col} '
                                                   f'Row: {row} Image: {img_file_name}')
            else:
                update_excel(write_cell_start, write_cell_end, url_cell_start, 'n/a', row, col)
                progress_recorder.set_progress(row, sh.max_row + 1,
                                               f'EMPTY CELL Column: {col} Row: {row}')
    
    wb.save(os.path.join(dir_path, xl_file_name))
    
    archive_and_save(archive_path, dir_path, file_url)
    return file_url


@shared_task(bind = True)
def process_split(self, split_value, xl_file_name):
    (cwd, table_location, media_path, dir_name, dir_path, folder, archive_file_name, archive_path,
     file_url) = get_paths(xl_file_name, 'split_files')
    
    wb = xl.load_workbook(table_location)
    sh = wb[wb.sheetnames[0]]
    
    build_table_location = platform_combine(folder, os.path.splitext(xl_file_name)[0], is_file = True)
    os.makedirs(folder)
    
    delete.delete_file_and_folders(media_path, 1)
    
    progress_recorder = ProgressRecorder(self)
    
    cut = split_value
    
    if sh.max_row >= split_value:
        while True:
            if cut == split_value:
                wb_build = xl.Workbook()
                sh_build = wb_build[wb_build.sheetnames[0]]
                for column_number in range(1, sh.max_column + 1):
                    for row_number in range(1, split_value + 1):
                        sh_build.cell(row_number, column_number).value = sh.cell(row_number, column_number).value
                        # progress_recorder.set_progress(row_number, sh.max_row + 1,
                        #                                f'Column: {column_number} Row: {row_number - 2}')
                wb_build.save(build_table_location + f" 1 - {cut}.xlsx")
                cut += split_value
            else:
                wb_build = xl.Workbook()
                sh_build = wb_build[wb_build.sheetnames[0]]
                
                if sh.max_row - (split_value - 1) >= cut - split_value:
                    for column_number in range(1, sh.max_column + 1):
                        sh_build.cell(1, column_number).value = sh.cell(1, column_number).value
                    wb_build.save(build_table_location + f" {cut - (split_value - 1)} - {cut}.xlsx")
                    
                    for column_number in range(1, sh.max_column + 1):
                        for row_number in range(cut - (split_value - 1), cut + 1):
                            # progress_recorder.set_progress(row_number, sh.max_row + 1,
                            #                                f'Column: {column_number} Row: {row_number - 2}')
                            sh_build.cell(row_number -
                                          (cut - (split_value + 1)), column_number).value = sh.cell(row_number,
                                                                                                    column_number).value
                    wb_build.save(build_table_location + f" {cut - (split_value - 1)} - {cut}.xlsx")
                    cut += split_value
                else:
                    wb_build = xl.Workbook()
                    sh_build = wb_build[wb_build.sheetnames[0]]
                    for column_number in range(1, sh.max_column + 1):
                        sh_build.cell(1, column_number).value = sh.cell(1, column_number).value
                    wb_build.save(build_table_location + f" {cut - (split_value - 1)} - {sh.max_row}.xlsx")
                    
                    for column_number in range(1, sh.max_column + 1):
                        for row_number in range(cut - (split_value - 1), sh.max_row + 1):
                            sh_build.cell(row_number - (cut - (split_value + 1)), column_number
                                          ).value = sh.cell(row_number, column_number).value
                            # progress_recorder.set_progress(row_number, sh.max_row + 1,
                            #                                f'Column: {column_number} Row: {row_number - 2}')
                    wb_build.save(build_table_location + f" {cut - (split_value - 1)} - {sh.max_row}.xlsx")
                    cut += split_value
                    break
    else:
        print('Table has less row than cut size')
    
    archive_and_save(archive_path, dir_path, file_url)
    
    return file_url


@shared_task(bind = True)
def keyv_process_file(self, f_key_column, l_value_column, f_empty_column, xl_file_name):
    (cwd, table_location, media_path, dir_name, dir_path, folder, archive_file_name, archive_path,
     file_url) = get_paths(xl_file_name, 'key_values')
    
    os.makedirs(folder)
    
    # delete.delete_file_and_folders(media_path, 1)
    
    progress_recorder = ProgressRecorder(self)
    
    wb = xl.load_workbook(table_location)
    sh = wb[wb.sheetnames[0]]
    
    keys = list(range(f_key_column, l_value_column + 1, 2))
    print('keys:', keys)
    values = list(range(f_key_column + 1, l_value_column + 1, 2))
    print('value:', values)
    col_dict = {}
    current_new_col = f_empty_column
    
    for row in range(2, sh.max_row + 1):
        progress_recorder.set_progress(row, sh.max_row + 1,
                                       f'Row: {row - 1}')
        for col_no in range(len(keys)):
            print(f'row: {row}, col: {keys[col_no]}')
            key = sh.cell(row, keys[col_no]).value
            if key is not None:
                if key in col_dict.keys():
                    col_val = col_dict[key]
                    sh.cell(row, col_val).value = sh.cell(row, values[col_no]).value
                else:
                    col_dict.update({key: current_new_col})
                    sh.cell(1, current_new_col).value = key
                    sh.cell(row, current_new_col).value = sh.cell(row, values[col_no]).value
                    current_new_col += 1
    
    wb.save(os.path.join(folder, xl_file_name))
    
    archive_and_save(archive_path, dir_path, file_url)
    
    return file_url
